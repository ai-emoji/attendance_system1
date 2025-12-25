"""export.export_grid_list

Xuất "lưới chấm công" từ QTableWidget (Shift Attendance - MainContent2) ra Excel.

Cấu trúc theo yêu cầu:
- Dòng 1: Tên công ty
- Dòng 2: Địa chỉ
- Dòng 3: Số điện thoại
- Dòng 4: Chi tiết chấm công
- Dòng 5: Từ ngày ... đến ngày ...
- Dòng 6: Header (chỉ các cột đang hiển thị)
- Dòng 7+: Dữ liệu trong bảng (đúng text đang hiển thị), không xuất cột ✅/❌.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class CompanyInfo:
    name: str = ""
    address: str = ""
    phone: str = ""


def export_shift_attendance_grid_xlsx(
    *,
    file_path: str,
    company: CompanyInfo,
    from_date_text: str,
    to_date_text: str,
    table,
    row_indexes: list[int] | None = None,
    force_exclude_headers: set[str] | None = None,
    company_name_style: dict | None = None,
    company_address_style: dict | None = None,
    company_phone_style: dict | None = None,
    creator: str | None = None,
    creator_style: dict | None = None,
    note_text: str | None = None,
    note_style: dict | None = None,
) -> tuple[bool, str]:
    """Export from a QTableWidget-like object to .xlsx.

    Parameters
    - file_path: output path (.xlsx)
    - company: CompanyInfo
    - from_date_text/to_date_text: already formatted for display (e.g. dd/MM/yyyy)
    - table: QTableWidget (or compatible API)
    """

    path = Path(str(file_path or "").strip())
    if not str(path):
        return False, "Vui lòng chọn đường dẫn file xuất."
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")

    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Border, Font, Side  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        return (
            False,
            "Thiếu thư viện xuất Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
        )

    # Determine visible columns (exclude checkbox column 0 and any hidden columns)
    excluded_headers = {str(x or "").strip() for x in (force_exclude_headers or set())}
    visible_cols: list[int] = []
    try:
        col_count = int(table.columnCount())
    except Exception:
        return False, "Không thể đọc cấu trúc cột của bảng để xuất."

    for c in range(col_count):
        if int(c) == 0:
            continue
        try:
            if bool(table.isColumnHidden(int(c))):
                continue
        except Exception:
            pass
        # Skip if header is empty (defensive)
        try:
            hi = table.horizontalHeaderItem(int(c))
            ht = "" if hi is None else str(hi.text() or "").strip()
        except Exception:
            ht = ""
        if not ht:
            continue
        if excluded_headers and ht in excluded_headers:
            continue
        visible_cols.append(int(c))

    if not visible_cols:
        return False, "Không có cột hiển thị để xuất (đang ẩn hết cột)."

    wb = Workbook()
    ws = wb.active
    ws.title = "XuatLuoi"

    ncols = len(visible_cols)

    def _merge_full_row(row: int) -> None:
        if ncols <= 1:
            return
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

    title_font = Font(bold=True)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _norm_style(
        st: dict | None, *, default_align: str = "center"
    ) -> tuple[Font, Alignment]:
        d = st or {}
        try:
            size = int(d.get("font_size", 13))
        except Exception:
            size = 13
        size = max(8, min(40, size))
        bold = bool(d.get("bold", False))
        italic = bool(d.get("italic", False))
        underline = bool(d.get("underline", False))
        align = str(d.get("align", default_align) or default_align).strip().lower()
        if align not in {"left", "center", "right"}:
            align = default_align
        f = Font(
            size=size,
            bold=bold,
            italic=italic,
            underline=("single" if underline else None),
        )
        a = Alignment(horizontal=align, vertical="center", wrap_text=True)
        return f, a

    # Row 1..5 (merged)
    _merge_full_row(1)
    c1 = ws.cell(
        row=1,
        column=1,
        value=f"Tên công ty : {str(company.name or '').strip()}",
    )
    if company_name_style is not None:
        f, a = _norm_style(company_name_style, default_align="center")
        c1.font = f
        c1.alignment = a
    else:
        c1.font = title_font
        c1.alignment = center

    _merge_full_row(2)
    c2 = ws.cell(
        row=2,
        column=1,
        value=f"Địa chỉ : {str(company.address or '').strip()}",
    )
    if company_address_style is not None:
        f, a = _norm_style(company_address_style, default_align="center")
        c2.font = f
        c2.alignment = a
    else:
        c2.alignment = center

    _merge_full_row(3)
    c3 = ws.cell(
        row=3,
        column=1,
        value=f"Số điện thoại : {str(company.phone or '').strip()}",
    )
    if company_phone_style is not None:
        f, a = _norm_style(company_phone_style, default_align="center")
        c3.font = f
        c3.alignment = a
    else:
        c3.alignment = center

    _merge_full_row(4)
    c4 = ws.cell(row=4, column=1, value="Chi tiết chấm công")
    c4.font = title_font
    c4.alignment = center

    _merge_full_row(5)
    c5 = ws.cell(
        row=5,
        column=1,
        value=f"Từ ngày: {str(from_date_text or '').strip()}    Đến ngày: {str(to_date_text or '').strip()}",
    )
    c5.font = title_font
    c5.alignment = center

    # Row 6: headers
    header_row = 6
    for excel_col, table_col in enumerate(visible_cols, start=1):
        try:
            hi = table.horizontalHeaderItem(int(table_col))
            label = "" if hi is None else str(hi.text() or "").strip()
        except Exception:
            label = ""
        cell = ws.cell(row=header_row, column=excel_col, value=label)
        cell.font = header_font
        cell.alignment = center

    # Data rows starting at row 7
    start_row = 7
    try:
        row_count = int(table.rowCount())
    except Exception:
        row_count = 0

    rows_to_export = list(range(row_count))
    if row_indexes is not None:
        rows_to_export = [
            int(r) for r in (row_indexes or []) if 0 <= int(r) < row_count
        ]

    for out_idx, r in enumerate(rows_to_export):
        excel_row = start_row + int(out_idx)
        for excel_col, table_col in enumerate(visible_cols, start=1):
            try:
                item = table.item(int(r), int(table_col))
                txt = "" if item is None else str(item.text() or "")
            except Exception:
                txt = ""
            cell = ws.cell(row=excel_row, column=excel_col, value=txt)
            cell.alignment = center

    # Full border for list area (header + data)
    thin = Side(style="thin")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    last_data_row = (start_row - 1) + len(rows_to_export)
    for r in range(header_row, max(header_row, last_data_row) + 1):
        for c in range(1, ncols + 1):
            try:
                ws.cell(row=int(r), column=int(c)).border = grid_border
            except Exception:
                pass

    # Basic widths: approximate from QTableWidget column width
    for excel_col, table_col in enumerate(visible_cols, start=1):
        try:
            px = int(table.columnWidth(int(table_col)))
        except Exception:
            px = 120
        # rough conversion px->excel width
        width = max(10, min(60, int(round(px / 7))))
        ws.column_dimensions[get_column_letter(excel_col)].width = width

    # Basic row heights for title area
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[6].height = 20

    # Append creator + note at bottom (ghi chú luôn ở cuối)
    creator_txt = str(creator or "").strip()
    note_txt = str(note_text or "").strip()
    if creator_txt or note_txt:
        base_row = last_data_row + 2

        def _merge_row(r: int) -> None:
            if ncols <= 1:
                return
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)

        if creator_txt:
            _merge_row(base_row)
            cs = creator_style or {}
            try:
                c_size = int(cs.get("font_size", 13))
            except Exception:
                c_size = 13
            c_size = max(8, min(40, c_size))
            c_bold = bool(cs.get("bold", False))
            c_italic = bool(cs.get("italic", False))
            c_underline = bool(cs.get("underline", False))
            c_align = str(cs.get("align", "left") or "left").strip().lower()
            if c_align not in {"left", "center", "right"}:
                c_align = "left"

            c_font = Font(
                size=c_size,
                bold=c_bold,
                italic=c_italic,
                underline=("single" if c_underline else None),
            )
            c_alignment = Alignment(
                horizontal=c_align,
                vertical="center",
                wrap_text=True,
            )
            c = ws.cell(
                row=base_row, column=1, value=f"Người tạo danh sách: {creator_txt}"
            )
            c.font = c_font
            c.alignment = c_alignment
            base_row += 1

        if note_txt:
            _merge_row(base_row)
            # Normalize style
            ns = note_style or {}
            try:
                size = int(ns.get("font_size", 13))
            except Exception:
                size = 13
            size = max(8, min(40, size))
            bold = bool(ns.get("bold", False))
            italic = bool(ns.get("italic", False))
            underline = bool(ns.get("underline", False))
            align = str(ns.get("align", "left") or "left").strip().lower()
            if align not in {"left", "center", "right"}:
                align = "left"

            note_font = Font(
                size=size,
                bold=bold,
                italic=italic,
                underline=("single" if underline else None),
            )
            note_align = Alignment(
                horizontal=align,
                vertical="top",
                wrap_text=True,
            )
            ncell = ws.cell(row=base_row, column=1, value=f"Ghi chú: {note_txt}")
            ncell.font = note_font
            ncell.alignment = note_align
            try:
                ws.row_dimensions[base_row].height = 40
            except Exception:
                pass

    # Ensure row heights are not too small (avoid clipping descenders).
    # Excel doesn't reliably auto-fit row height for wrapped/merged cells on open,
    # so we enforce a sensible minimum and scale with font size.
    try:
        min_height = 30
        max_row = int(ws.max_row or 0)
        for r in range(1, max_row + 1):
            current_h = ws.row_dimensions[int(r)].height

            max_font_size = 11
            for c in range(1, ncols + 1):
                try:
                    cell = ws.cell(row=int(r), column=int(c))
                    fs = getattr(getattr(cell, "font", None), "size", None)
                    if fs is None:
                        continue
                    max_font_size = max(max_font_size, int(fs))
                except Exception:
                    pass

            # Rough conversion: points ~= font_size*1.4 + padding
            target_h = max(min_height, int(round(max_font_size * 1.4 + 8)))
            if current_h is None or float(current_h) < float(target_h):
                ws.row_dimensions[int(r)].height = target_h
    except Exception:
        pass

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return True, f"Đã xuất dữ liệu: {path}"
