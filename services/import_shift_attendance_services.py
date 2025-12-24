"""services.import_shift_attendance_services

Service cho tính năng "Import dữ liệu chấm công" -> ghi vào attendance_audit.

Luồng:
- export_shift_attendance_template_xlsx: tạo file mẫu
- read_shift_attendance_from_xlsx: đọc Excel -> list rows để preview
- import_shift_attendance_rows: áp dụng vào DB với rules overwrite/skip

Rules overwrite/skip (theo yêu cầu):
- Nếu dòng audit hiện có có import_locked = 0 (nguồn download/sync): overwrite luôn.
- Nếu import_locked = 1 (đã từng import): so sánh tất cả field -> chỉ update khi có thay đổi, không đổi thì SKIP.

Ghi chú:
- File Excel mẫu/preview theo đúng cột MainContent2 (không có attendance_code/device_no).
- Khi import: nếu đã có dữ liệu audit theo (employee_code, work_date) thì dùng (attendance_code, device_no) hiện có để upsert.
- Nếu chưa có: sẽ cố gắng map attendance_code = employees.mcc_code (nếu có) else employee_code; device_no mặc định = 1.
"""

from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable

from repository.import_shift_attendance_repository import (
    ImportShiftAttendanceRepository,
)


logger = logging.getLogger(__name__)


@dataclass
class ImportShiftAttendanceResult:
    ok: bool
    message: str
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    failed: int = 0


class ImportShiftAttendanceService:
    def __init__(
        self, repository: ImportShiftAttendanceRepository | None = None
    ) -> None:
        self._repo = repository or ImportShiftAttendanceRepository()

    @staticmethod
    def _weekday_label(d: date) -> str:
        # 0=Mon..6=Sun
        w = int(d.weekday())
        return (
            "Thứ 2"
            if w == 0
            else (
                "Thứ 3"
                if w == 1
                else (
                    "Thứ 4"
                    if w == 2
                    else (
                        "Thứ 5"
                        if w == 3
                        else "Thứ 6" if w == 4 else "Thứ 7" if w == 5 else "Chủ nhật"
                    )
                )
            )
        )

    @staticmethod
    def export_shift_attendance_template_xlsx(file_path: str) -> tuple[bool, str]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng chọn đường dẫn lưu file mẫu."
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")

        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font  # type: ignore
        except Exception:
            return (
                False,
                "Thiếu thư viện ghi Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
            )

        headers = [
            "Mã nv",
            "Tên nhân viên",
            "Ngày",
            "Thứ",
            "Vào 1",
            "Ra 1",
            "Vào 2",
            "Ra 2",
            "Vào 3",
            "Ra 3",
            "Trễ",
            "Sớm",
            "Giờ",
            "Công",
            "KH",
            "Giờ +",
            "Công +",
            "KH +",
            "TC1",
            "TC2",
            "TC3",
            "Lịch NV",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "ChamCong"

        header_font = Font(bold=True)
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font

        for col_idx, label in enumerate(headers, start=1):
            width = max(12, min(28, len(label) + 6))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                width
            )

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return True, f"Đã tạo file mẫu: {path}"

    def read_shift_attendance_from_xlsx(
        self, file_path: str
    ) -> tuple[bool, str, list[dict[str, Any]]]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng nhập đường dẫn file Excel.", []
        if not path.exists() or path.suffix.lower() != ".xlsx":
            return False, "Vui lòng chọn file .xlsx hợp lệ.", []

        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            return (
                False,
                "Thiếu thư viện đọc Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
                [],
            )

        def norm_header(s: Any) -> str:
            raw = str(s or "").strip()
            raw = re.sub(r"\s+", " ", raw)
            return raw

        def norm_header_key(s: Any) -> str:
            raw = str(s or "").strip().lower()
            raw = re.sub(r"\s+", " ", raw)
            raw = raw.replace("đ", "d").replace("Đ", "d")
            raw = raw.replace("\ufffd", "")
            raw = unicodedata.normalize("NFKD", raw)
            raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
            raw = re.sub(r"[^0-9a-z ]+", "", raw)
            raw = raw.replace(" ", "")
            return raw

        header_to_key: dict[str, str] = {
            # keys
            "employee_code": "employee_code",
            "full_name": "full_name",
            "date": "work_date",
            "work_date": "work_date",
            "weekday": "weekday",
            "schedule": "schedule",
            "in_1": "in_1",
            "out_1": "out_1",
            "in_2": "in_2",
            "out_2": "out_2",
            "in_3": "in_3",
            "out_3": "out_3",
            "late": "late",
            "early": "early",
            "hours": "hours",
            "work": "work",
            "leave": "leave",
            "hours_plus": "hours_plus",
            "work_plus": "work_plus",
            "leave_plus": "leave_plus",
            "tc1": "tc1",
            "tc2": "tc2",
            "tc3": "tc3",
            # Vietnamese (MainContent2)
            "Mã nv": "employee_code",
            "Mã NV": "employee_code",
            "Tên nhân viên": "full_name",
            "Họ và tên": "full_name",
            "Ngày": "work_date",
            "Thứ": "weekday",
            "Lịch NV": "schedule",
            "Vào 1": "in_1",
            "Ra 1": "out_1",
            "Vào 2": "in_2",
            "Ra 2": "out_2",
            "Vào 3": "in_3",
            "Ra 3": "out_3",
            "Trễ": "late",
            "Sớm": "early",
            "Giờ": "hours",
            "Công": "work",
            "KH": "leave",
            "Giờ +": "hours_plus",
            "Công +": "work_plus",
            "KH +": "leave_plus",
            "TC1": "tc1",
            "TC2": "tc2",
            "TC3": "tc3",
            # Backward-compat: older template used 'Tổng' column
            "Tổng": "schedule",
            # ignored
            "STT": "stt",
            "ID": "id",
            "__check": "__check",
            "": "",
        }

        def parse_date(v: Any) -> str | None:
            if v is None:
                return None
            if isinstance(v, datetime):
                return v.date().isoformat()
            if isinstance(v, date):
                return v.isoformat()
            s = str(v or "").strip()
            if not s:
                return None
            # dd/MM/yyyy
            try:
                if "/" in s and len(s.split("/")) == 3:
                    dd, mm, yy = s.split("/")
                    return date(int(yy), int(mm), int(dd)).isoformat()
            except Exception:
                pass
            try:
                return date.fromisoformat(s).isoformat()
            except Exception:
                return None

        def parse_time(v: Any) -> time | None:
            if v is None:
                return None
            if isinstance(v, time):
                return v
            if isinstance(v, datetime):
                return v.time().replace(microsecond=0)
            s = str(v or "").strip()
            if not s:
                return None
            s = s.replace(".", ":")
            parts = s.split(":")
            try:
                if len(parts) == 2:
                    hh, mm = parts
                    return time(int(hh), int(mm), 0)
                if len(parts) >= 3:
                    hh, mm, ss = parts[:3]
                    return time(int(hh), int(mm), int(float(ss)))
            except Exception:
                return None
            return None

        def parse_decimal(v: Any) -> Decimal | None:
            if v is None:
                return None
            if isinstance(v, Decimal):
                return v
            if isinstance(v, (int, float)):
                try:
                    return Decimal(str(v))
                except Exception:
                    return None
            s = str(v or "").strip()
            if not s:
                return None
            s = s.replace(" ", "")
            # Vietnamese style: 1,5
            if s.count(",") == 1 and s.count(".") == 0:
                s = s.replace(",", ".")
            try:
                return Decimal(s)
            except Exception:
                return None

        wb = load_workbook(str(path), data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return False, "File Excel trống.", []

        headers = [norm_header(h) for h in list(header_row or [])]
        header_to_key_lower = {
            str(k).strip().lower(): v for k, v in header_to_key.items()
        }
        header_to_key_norm = {norm_header_key(k): v for k, v in header_to_key.items()}

        col_keys: list[str | None] = []
        for h in headers:
            key = header_to_key.get(h)
            if key is None:
                key = header_to_key_lower.get(str(h or "").strip().lower())
            if key is None:
                key = header_to_key_norm.get(norm_header_key(h))
            col_keys.append(key)

        unknown_headers = [
            headers[i]
            for i, k in enumerate(col_keys)
            if k is None and str(headers[i] or "").strip()
        ]
        if unknown_headers:
            # Không fail; chỉ cảnh báo nhẹ trong message
            pass

        out: list[dict[str, Any]] = []
        for r in rows_iter:
            if r is None:
                continue
            item: dict[str, Any] = {}
            empty = True
            for idx, raw in enumerate(list(r)):
                key = col_keys[idx] if idx < len(col_keys) else None
                if not key or key in {"id", "stt", "__check"}:
                    continue

                if raw is not None and str(raw).strip() != "":
                    empty = False

                if key == "work_date":
                    item[key] = parse_date(raw)
                elif key in {"in_1", "out_1", "in_2", "out_2", "in_3", "out_3"}:
                    item[key] = parse_time(raw)
                elif key in {
                    "hours",
                    "work",
                    "leave",
                    "hours_plus",
                    "work_plus",
                    "leave_plus",
                }:
                    item[key] = parse_decimal(raw)
                else:
                    s = str(raw or "").strip()
                    item[key] = s if s else None

            if empty:
                continue

            emp_code = str(item.get("employee_code") or "").strip()
            if emp_code:
                item["employee_code"] = emp_code

            wd = str(item.get("work_date") or "").strip()
            if wd:
                item["work_date"] = wd
                try:
                    d = date.fromisoformat(wd)
                    item.setdefault("weekday", self._weekday_label(d))
                except Exception:
                    pass

            out.append(item)

        msg = f"Đọc thành công {len(out)} dòng."
        if unknown_headers:
            msg += f" (Bỏ qua cột lạ: {', '.join(unknown_headers[:6])}{'...' if len(unknown_headers) > 6 else ''})"
        return True, msg, out

    def import_shift_attendance_rows(
        self,
        rows: list[dict[str, Any]],
        progress_cb: Callable[[int, bool, str, str], None] | None = None,
        report: list[dict[str, Any]] | None = None,
    ) -> ImportShiftAttendanceResult:
        if not rows:
            return ImportShiftAttendanceResult(False, "Không có dữ liệu để cập nhập.")

        # Build keys for existing lookup
        pairs: list[tuple[str, str]] = []
        emp_codes: list[str] = []
        for r in rows:
            emp_code = str(r.get("employee_code") or "").strip()
            wd = str(r.get("work_date") or "").strip()
            if emp_code and wd:
                pairs.append((emp_code, wd))
                emp_codes.append(emp_code)

        existing_map = {}
        try:
            existing_map = self._repo.get_existing_by_employee_code_date(pairs)
        except Exception:
            # if DB lookup fails, we still allow insert best-effort
            existing_map = {}

        # Employee lookup (for attendance_code + employee_id)
        emp_lookup: dict[str, dict[str, Any]] = {}
        try:
            emp_lookup = self._repo.get_employees_by_codes(emp_codes)
        except Exception:
            emp_lookup = {}

        def to_time_str(t: Any) -> str | None:
            if t is None:
                return None
            if isinstance(t, time):
                return t.replace(microsecond=0).strftime("%H:%M:%S")
            return str(t)

        def to_dec_str(v: Any) -> str | None:
            if v is None:
                return None
            if isinstance(v, Decimal):
                return format(v, "f")
            try:
                return format(Decimal(str(v)), "f")
            except Exception:
                s = str(v or "").strip()
                return s if s else None

        compare_keys = [
            "employee_code",
            "full_name",
            "work_date",
            "weekday",
            "schedule",
            "in_1",
            "out_1",
            "in_2",
            "out_2",
            "in_3",
            "out_3",
            "late",
            "early",
            "hours",
            "work",
            "leave",
            "hours_plus",
            "work_plus",
            "leave_plus",
            "tc1",
            "tc2",
            "tc3",
        ]

        def normalize_for_compare(k: str, v: Any) -> Any:
            if k in {"in_1", "out_1", "in_2", "out_2", "in_3", "out_3"}:
                return to_time_str(v)
            if k in {
                "hours",
                "work",
                "leave",
                "hours_plus",
                "work_plus",
                "leave_plus",
            }:
                return to_dec_str(v)
            if k == "work_date":
                if v is None:
                    return None
                if isinstance(v, (date, datetime)):
                    return (v.date() if isinstance(v, datetime) else v).isoformat()
                return str(v)
            s = str(v or "").strip()
            return s if s else None

        def add_report(
            *,
            idx: int,
            code: str,
            result: str,
            action: str,
            message: str,
        ) -> None:
            if report is None:
                return
            report.append(
                {
                    "index": int(idx),
                    "employee_code": str(code or "").strip(),
                    "result": str(result or ""),
                    "action": str(action or ""),
                    "message": str(message or ""),
                }
            )

        upsert_payloads: list[dict[str, Any]] = []

        inserted = 0
        updated = 0
        skipped = 0
        failed = 0

        total = len(rows)
        for i, raw in enumerate(rows, start=1):
            emp_code = str(raw.get("employee_code") or "").strip()
            wd = str(raw.get("work_date") or "").strip()

            if not emp_code or not wd:
                failed += 1
                add_report(
                    idx=i,
                    code=emp_code or "(không mã)",
                    result="INVALID",
                    action="SKIP_INVALID",
                    message="Thiếu Mã nv hoặc Ngày.",
                )
                if progress_cb:
                    progress_cb(i, False, emp_code or "(không mã)", "Thiếu dữ liệu")
                continue

            existing = existing_map.get((emp_code, wd))
            import_locked = int(existing.get("import_locked") or 0) if existing else 0

            # Determine whether changed (only when import_locked=1)
            changed = True
            if existing and import_locked == 1:
                changed = False
                for k in compare_keys:
                    new_v = normalize_for_compare(k, raw.get(k))
                    old_v = normalize_for_compare(k, existing.get(k))
                    if new_v != old_v:
                        changed = True
                        break

            if existing and import_locked == 1 and not changed:
                skipped += 1
                add_report(
                    idx=i,
                    code=emp_code,
                    result="SKIPPED",
                    action="SKIP_NO_CHANGE",
                    message="Không thay đổi.",
                )
                if progress_cb:
                    progress_cb(i, True, emp_code, "Bỏ qua (không đổi)")
                continue

            # Build payload for upsert
            payload: dict[str, Any] = {}

            # Resolve base keys for unique
            if existing:
                payload["attendance_code"] = (
                    str(existing.get("attendance_code") or "").strip() or emp_code
                )
                payload["device_no"] = int(existing.get("device_no") or 1)
                payload["device_id"] = existing.get("device_id")
                payload["device_name"] = existing.get("device_name")
            else:
                emp = emp_lookup.get(emp_code.lower())
                mcc = str((emp or {}).get("mcc_code") or "").strip()
                payload["attendance_code"] = mcc or emp_code
                payload["device_no"] = 1
                payload["device_id"] = None
                payload["device_name"] = ""

            emp = emp_lookup.get(emp_code.lower())
            payload["employee_id"] = (
                int(emp.get("id")) if emp and emp.get("id") is not None else None
            )
            payload["employee_code"] = emp_code

            name = str(raw.get("full_name") or "").strip()
            if not name and emp:
                name = str(emp.get("full_name") or emp.get("name_on_mcc") or "").strip()
            payload["full_name"] = name or None

            payload["work_date"] = wd
            # weekday: use provided, else compute
            weekday = str(raw.get("weekday") or "").strip()
            if not weekday:
                try:
                    weekday = self._weekday_label(date.fromisoformat(wd))
                except Exception:
                    weekday = ""
            payload["weekday"] = weekday or None

            # Times + fields
            for k in ["in_1", "out_1", "in_2", "out_2", "in_3", "out_3"]:
                payload[k] = raw.get(k)

            for k in [
                "late",
                "early",
                "tc1",
                "tc2",
                "tc3",
            ]:
                v = str(raw.get(k) or "").strip()
                payload[k] = v or None

            for k in [
                "hours",
                "work",
                "leave",
                "hours_plus",
                "work_plus",
                "leave_plus",
            ]:
                payload[k] = raw.get(k)

            # schedule is a display field (varchar)
            schedule = str(raw.get("schedule") or "").strip()
            payload["schedule"] = schedule or None

            # Mark as imported
            payload["import_locked"] = 1

            # Decide action label
            if existing:
                action = (
                    "OVERWRITE_DOWNLOAD" if import_locked == 0 else "UPDATE_CHANGED"
                )
                updated += 1
            else:
                action = "INSERT"
                inserted += 1

            upsert_payloads.append(payload)
            add_report(
                idx=i,
                code=emp_code,
                result="SUCCESS",
                action=action,
                message="Sẽ cập nhập" if existing else "Sẽ thêm",
            )
            if progress_cb:
                progress_cb(i, True, emp_code, action)

        # Execute upserts in one batch
        try:
            self._repo.upsert_import_rows(upsert_payloads)
        except Exception as exc:
            # Mark remaining as failed (best-effort)
            logger.exception("Import attendance_audit thất bại")
            return ImportShiftAttendanceResult(
                False,
                f"Không thể cập nhập CSDL: {exc}",
                inserted=inserted,
                updated=updated,
                skipped=skipped,
                failed=(failed + max(0, len(upsert_payloads))),
            )

        msg = f"Hoàn tất import: Thêm {inserted}, Cập nhập {updated}, Bỏ qua {skipped}, Lỗi {failed}."
        return ImportShiftAttendanceResult(
            True,
            msg,
            inserted=inserted,
            updated=updated,
            skipped=skipped,
            failed=failed,
        )
