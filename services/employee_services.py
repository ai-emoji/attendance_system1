"""services.employee_services

Service cho màn Thông tin Nhân viên:
- list employees theo filter
- export/import CSV
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Any
from typing import Callable
import re
import unicodedata

from repository.employee_repository import EmployeeRepository
from repository.schedule_work_repository import ScheduleWorkRepository
from services.department_services import DepartmentService
from services.title_services import TitleService


class EmployeeService:
    @staticmethod
    def _parse_bool(v: Any) -> bool | None:
        """Parse cell value into boolean.

        Supports:
        - bool
        - numbers (0/1, including 1.0)
        - strings: 1/0, true/false, yes/no, x, có/không, and strings containing numbers (e.g. '01 năm')
        """

        if v is None:
            return None
        if isinstance(v, bool):
            return bool(v)

        if isinstance(v, (int, float)):
            try:
                return float(v) != 0.0
            except Exception:
                return None

        s = str(v or "").strip().lower()
        if not s:
            return None

        # numeric-looking strings like "1.0"
        try:
            return float(s) != 0.0
        except Exception:
            pass

        # strings like "01 năm", "02 nam"...
        m = re.search(r"(-?\d+(?:[\.,]\d+)?)", s)
        if m:
            num = m.group(1).replace(",", ".")
            try:
                return float(num) != 0.0
            except Exception:
                pass

        true_set = {"1", "true", "yes", "y", "x", "có", "co", "✓", "✔"}
        false_set = {"0", "false", "no", "n", "không", "khong"}
        if s in true_set:
            return True
        if s in false_set:
            return False
        return None

    @staticmethod
    def _parse_date_for_db(v: Any) -> str | None:
        """Parse input into ISO date string (YYYY-MM-DD) for DB."""

        if v is None:
            return None

        # openpyxl returns datetime/date objects
        try:
            if hasattr(v, "date"):
                d = v.date() if hasattr(v, "hour") else v
                return d.isoformat()
        except Exception:
            pass

        s = str(v or "").strip()
        if not s:
            return None

        # Full date formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.date().isoformat()
            except Exception:
                continue

        # Month/year formats like 07/2013 or 07-2013
        m = re.match(r"^(\d{1,2})[\/-](\d{4})$", s)
        if m:
            try:
                mm = int(m.group(1))
                yy = int(m.group(2))
                if 1 <= mm <= 12:
                    return f"{yy:04d}-{mm:02d}-01"
            except Exception:
                return None

        return None

    @staticmethod
    def _date_value_for_preview(v: Any) -> Any:
        """Keep Excel date display as-is when it is a string (e.g. '07/2013').

        For true date/datetime objects, store ISO (table model will format nicely).
        """

        if v is None:
            return None
        try:
            if hasattr(v, "date"):
                d = v.date() if hasattr(v, "hour") else v
                return d.isoformat()
        except Exception:
            pass

        s = str(v or "").strip()
        return s if s != "" else None

    def __init__(
        self,
        repo: EmployeeRepository | None = None,
        department_service: DepartmentService | None = None,
        title_service: TitleService | None = None,
        schedule_work_repo: ScheduleWorkRepository | None = None,
    ) -> None:
        self._repo = repo or EmployeeRepository()
        self._department_service = department_service or DepartmentService()
        self._title_service = title_service or TitleService()
        self._schedule_work_repo = schedule_work_repo or ScheduleWorkRepository()

    def list_departments_tree_rows(self) -> list[tuple[int, int | None, str, str]]:
        models = self._department_service.list_departments()
        return [
            (m.id, m.parent_id, m.department_name, m.department_note) for m in models
        ]

    def list_employees(self, filters: dict) -> list[dict[str, Any]]:
        # Backward compatible: allow both old (employee_code/full_name) and new
        # (search_by/search_text) filter shapes.
        search_by = str(filters.get("search_by") or "").strip() or None
        search_text = str(filters.get("search_text") or "").strip() or None
        employment_status = str(filters.get("employment_status") or "").strip() or None

        # Legacy fields fallback
        legacy_code = str(filters.get("employee_code") or "").strip() or None
        legacy_name = str(filters.get("full_name") or "").strip() or None

        employee_code: str | None = None
        full_name: str | None = None
        mcc_code: str | None = None
        sort_order: int | None = None

        if search_text and search_by:
            if search_by == "stt":
                try:
                    sort_order = int(float(search_text))
                except Exception:
                    sort_order = None
            elif search_by == "full_name":
                full_name = search_text
            elif search_by == "mcc_code":
                mcc_code = search_text
            else:
                employee_code = search_text
        else:
            employee_code = legacy_code
            full_name = legacy_name

        rows = self._repo.list_employees(
            employee_code=employee_code,
            mcc_code=mcc_code,
            full_name=full_name,
            sort_order=sort_order,
            employment_status=employment_status,
            department_id=filters.get("department_id"),
            title_id=filters.get("title_id"),
        )

        # Gắn thêm "lịch làm việc" (tên lịch) cho từng nhân viên theo ngày hiện tại.
        # Không ảnh hưởng các màn hình không hiển thị cột này.
        try:
            emp_ids: list[int] = []
            for r in rows or []:
                try:
                    v = int(r.get("id") or 0)
                except Exception:
                    v = 0
                if v > 0:
                    emp_ids.append(v)
            emp_ids = list(dict.fromkeys(emp_ids))

            schedule_map = self._schedule_work_repo.get_employee_schedule_name_map(
                employee_ids=emp_ids,
                on_date=date.today().isoformat(),
            )
            for r in rows or []:
                try:
                    eid = int(r.get("id") or 0)
                except Exception:
                    eid = 0
                r["schedule"] = str(schedule_map.get(eid) or "")
        except Exception:
            # Best-effort: nếu lỗi DB/thiếu bảng thì vẫn trả danh sách nhân viên bình thường
            for r in rows or []:
                r["schedule"] = ""

        return rows

    def list_departments_dropdown(self) -> list[tuple[int, str]]:
        models = self._department_service.list_departments()
        items: list[tuple[int, str]] = []
        for m in models:
            try:
                items.append((int(m.id), str(m.department_name)))
            except Exception:
                continue
        items.sort(key=lambda x: x[1].lower())
        return items

    def list_titles_dropdown(self) -> list[tuple[int, str]]:
        models = self._title_service.list_titles()
        items: list[tuple[int, str]] = []
        for m in models:
            try:
                items.append((int(m.id), str(m.title_name)))
            except Exception:
                continue
        items.sort(key=lambda x: x[1].lower())
        return items

    def list_issue_places_dropdown(self) -> list[str]:
        try:
            return self._repo.list_distinct_id_issue_places()
        except Exception:
            return []

    def get_employee(self, employee_id: int) -> dict[str, Any] | None:
        row = self._repo.get_employee(int(employee_id))
        if not row:
            return None

        def to_str(v: Any) -> Any:
            if v is None:
                return None
            try:
                return v.isoformat()
            except Exception:
                return v

        # Normalize dates to ISO strings for UI
        for k in (
            "start_date",
            "date_of_birth",
            "id_issue_date",
            "contract1_sign_date",
            "contract1_expire_date",
            "contract2_sign_date",
            "child_dob_1",
            "child_dob_2",
            "child_dob_3",
            "child_dob_4",
        ):
            row[k] = to_str(row.get(k))
        return row

    def export_csv(self, file_path: str, filters: dict) -> tuple[bool, str]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng chọn đường dẫn file xuất."
        if path.suffix.lower() != ".csv":
            path = path.with_suffix(".csv")

        rows = self.list_employees(filters)

        headers = [
            "id",
            "employee_code",
            "mcc_code",
            "full_name",
            "name_on_mcc",
            "start_date",
            "title_name",
            "department_name",
            "date_of_birth",
            "gender",
            "national_id",
            "id_issue_date",
            "id_issue_place",
            "address",
            "phone",
            "insurance_no",
            "tax_code",
            "degree",
            "major",
            "contract1_signed",
            "contract1_no",
            "contract1_sign_date",
            "contract1_expire_date",
            "contract2_indefinite",
            "contract2_no",
            "contract2_sign_date",
            "children_count",
            "child_dob_1",
            "child_dob_2",
            "child_dob_3",
            "child_dob_4",
            "employment_status",
            "note",
        ]

        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader()
            for r in rows:
                w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in headers})

        return True, f"Đã xuất {len(rows)} dòng: {path}"

    def export_xlsx(self, file_path: str, filters: dict) -> tuple[bool, str]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng chọn đường dẫn file xuất."
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")

        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font  # type: ignore
        except Exception:
            return (
                False,
                "Thiếu thư viện xuất Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
            )

        rows = self.list_employees(filters)

        return self.export_xlsx_rows(str(path), rows)

    def export_xlsx_rows(
        self, file_path: str, rows: list[dict[str, Any]]
    ) -> tuple[bool, str]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng chọn đường dẫn file xuất."
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")

        columns: list[tuple[str, str]] = [
            ("STT", "stt"),
            ("Mã NV", "employee_code"),
            ("Mã CC", "mcc_code"),
            ("Họ và tên", "full_name"),
            ("Tên trên MCC", "name_on_mcc"),
            ("Ngày vào làm", "start_date"),
            ("Chức Vụ", "title_name"),
            ("Phòng Ban", "department_name"),
            ("Ngày tháng năm sinh", "date_of_birth"),
            ("Giới tính", "gender"),
            ("CCCD/CMT", "national_id"),
            ("Ngày Cấp", "id_issue_date"),
            ("Nơi Cấp", "id_issue_place"),
            ("Địa chỉ", "address"),
            ("Số điện thoại", "phone"),
            ("Số Bảo Hiểm", "insurance_no"),
            ("Mã số Thuế TNCN", "tax_code"),
            ("Bằng cấp", "degree"),
            ("Chuyên ngành", "major"),
            ("HĐLĐ (ký lần 1)", "contract1_signed"),
            ("Số HĐLĐ (lần 1)", "contract1_no"),
            ("Ngày ký (lần 1)", "contract1_sign_date"),
            ("Ngày hết hạn (lần 1)", "contract1_expire_date"),
            ("HĐLĐ ký không thời hạn", "contract2_indefinite"),
            ("Số HĐLĐ (không thời hạn)", "contract2_no"),
            ("Ngày ký (không thời hạn)", "contract2_sign_date"),
            ("Số con", "children_count"),
            ("Ngày sinh con 1", "child_dob_1"),
            ("Ngày sinh con 2", "child_dob_2"),
            ("Ngày sinh con 3", "child_dob_3"),
            ("Ngày sinh con 4", "child_dob_4"),
            ("Hiện trạng", "employment_status"),
            ("Ghi chú", "note"),
        ]

        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font  # type: ignore
        except Exception:
            return (
                False,
                "Thiếu thư viện xuất Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "NhanVien"

        header_font = Font(bold=True)
        for col_idx, (label, _key) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font

        for row_idx, r in enumerate(rows, start=2):
            for col_idx, (_label, key) in enumerate(columns, start=1):
                v = r.get(key)
                if key in {"contract1_signed", "contract2_indefinite"}:
                    v = "1" if bool(v) else "0"
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=("" if v is None else v),
                )

        # basic width
        for col_idx, (label, _key) in enumerate(columns, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                max(12, min(40, len(label) + 6))
            )

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return True, f"Đã xuất {len(rows)} dòng: {path}"

    def export_employee_template_xlsx(self, file_path: str) -> tuple[bool, str]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng chọn đường dẫn file mẫu."
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")

        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font  # type: ignore
        except Exception:
            return (
                False,
                "Thiếu thư viện đọc/ghi Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
            )

        # Use Vietnamese headers (same spirit as export_xlsx) so users can fill in easily.
        columns: list[tuple[str, str]] = [
            ("STT", "stt"),
            ("MÃ NV", "employee_code"),
            ("Mã CC", "mcc_code"),
            ("HỌ VÀ TÊN", "full_name"),
            ("Tên trên MCC", "name_on_mcc"),
            ("Ngày vào làm", "start_date"),
            ("Chức Vụ", "title_name"),
            ("Phòng Ban", "department_name"),
            ("Ngày tháng năm sinh", "date_of_birth"),
            ("Giới tính", "gender"),
            ("CCCD/CMT", "national_id"),
            ("Ngày Cấp", "id_issue_date"),
            ("Nơi Cấp", "id_issue_place"),
            ("Địa chỉ", "address"),
            ("Số điện thoại", "phone"),
            ("Số Bảo Hiểm", "insurance_no"),
            ("Mã số Thuế TNCN", "tax_code"),
            ("Bằng cấp", "degree"),
            ("Chuyên ngành", "major"),
            ("HĐLĐ (ký lần 1)", "contract1_signed"),
            ("Số HĐLĐ (lần 1)", "contract1_no"),
            ("Ngày ký (lần 1)", "contract1_sign_date"),
            ("Ngày hết hạn (lần 1)", "contract1_expire_date"),
            ("HĐLĐ ký không thời hạn", "contract2_indefinite"),
            ("Số HĐLĐ (không thời hạn)", "contract2_no"),
            ("Ngày ký (không thời hạn)", "contract2_sign_date"),
            ("Số con", "children_count"),
            ("Ngày sinh con 1", "child_dob_1"),
            ("Ngày sinh con 2", "child_dob_2"),
            ("Ngày sinh con 3", "child_dob_3"),
            ("Ngày sinh con 4", "child_dob_4"),
            ("Hiện trạng", "employment_status"),
            ("Ghi chú", "note"),
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "NhanVien"

        header_font = Font(bold=True)
        for col_idx, (label, _key) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font

        for col_idx, (label, _key) in enumerate(columns, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                max(12, min(40, len(label) + 6))
            )

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return True, f"Đã tạo file mẫu: {path}"

    def read_employees_from_xlsx(
        self, file_path: str
    ) -> tuple[bool, str, list[dict[str, Any]]]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng nhập đường dẫn file Excel.", []
        if not path.exists() or path.suffix.lower() != ".xlsx":
            return False, "Vui lòng chọn file .xlsx hợp lệ.", []

        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            return (
                False,
                "Thiếu thư viện đọc Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
                [],
            )

        def norm_header(s: Any) -> str:
            # Normalize headers so Excel variations like "Phòng ban" vs "Phòng Ban"
            # or double spaces still map correctly.
            raw = str(s or "").strip()
            raw = re.sub(r"\s+", " ", raw)
            return raw

        def norm_header_key(s: Any) -> str:
            # Robust matching for Vietnamese headers even if the file contains
            # replacement characters ("�") or different diacritics.
            raw = str(s or "").strip().lower()
            raw = re.sub(r"\s+", " ", raw)
            raw = raw.replace("đ", "d").replace("Đ", "d")
            raw = raw.replace("\ufffd", "")  # Unicode replacement char
            raw = unicodedata.normalize("NFKD", raw)
            raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
            raw = re.sub(r"[^0-9a-z ]+", "", raw)
            raw = raw.replace(" ", "")
            return raw

        header_to_key: dict[str, str] = {
            # keys
            "employee_code": "employee_code",
            "mcc_code": "mcc_code",
            "full_name": "full_name",
            "name_on_mcc": "name_on_mcc",
            "start_date": "start_date",
            "title_name": "title_name",
            "department_name": "department_name",
            "date_of_birth": "date_of_birth",
            "gender": "gender",
            "national_id": "national_id",
            "id_issue_date": "id_issue_date",
            "id_issue_place": "id_issue_place",
            "address": "address",
            "phone": "phone",
            "insurance_no": "insurance_no",
            "tax_code": "tax_code",
            "degree": "degree",
            "major": "major",
            "contract1_signed": "contract1_signed",
            "contract1_no": "contract1_no",
            "contract1_sign_date": "contract1_sign_date",
            "contract1_expire_date": "contract1_expire_date",
            "contract2_indefinite": "contract2_indefinite",
            "contract2_no": "contract2_no",
            "contract2_sign_date": "contract2_sign_date",
            "children_count": "children_count",
            "child_dob_1": "child_dob_1",
            "child_dob_2": "child_dob_2",
            "child_dob_3": "child_dob_3",
            "child_dob_4": "child_dob_4",
            "employment_status": "employment_status",
            "note": "note",
            # Vietnamese
            "MÃ NV": "employee_code",
            "Mã NV": "employee_code",
            "Mã CC": "mcc_code",
            "MÃ CC": "mcc_code",
            "HỌ VÀ TÊN": "full_name",
            "Họ và tên": "full_name",
            "Tên trên MCC": "name_on_mcc",
            "TÊN TRÊN MCC": "name_on_mcc",
            "Ngày vào làm": "start_date",
            "Chức Vụ": "title_name",
            "Phòng Ban": "department_name",
            "Ngày tháng năm sinh": "date_of_birth",
            "Giới tính": "gender",
            "CCCD/CMT": "national_id",
            "Ngày Cấp": "id_issue_date",
            "Nơi Cấp": "id_issue_place",
            "Địa chỉ": "address",
            "Số điện thoại": "phone",
            "Số Bảo Hiểm": "insurance_no",
            "Mã số Thuế TNCN": "tax_code",
            "Bằng cấp": "degree",
            "Chuyên ngành": "major",
            "HĐLĐ (ký lần 1)": "contract1_signed",
            "Số HĐLĐ (lần 1)": "contract1_no",
            "Ngày ký (lần 1)": "contract1_sign_date",
            "Ngày hết hạn (lần 1)": "contract1_expire_date",
            "HĐLĐ ký không thời hạn": "contract2_indefinite",
            "Số HĐLĐ (không thời hạn)": "contract2_no",
            "Ngày ký (không thời hạn)": "contract2_sign_date",
            "Số con": "children_count",
            "Ngày sinh con 1": "child_dob_1",
            "Ngày sinh con 2": "child_dob_2",
            "Ngày sinh con 3": "child_dob_3",
            "Ngày sinh con 4": "child_dob_4",
            "Hiện trạng": "employment_status",
            "Hien trang": "employment_status",
            "Ghi chú": "note",
            # ignored
            "STT": "stt",
            "ID": "id",
        }

        def parse_int(v: Any) -> int | None:
            if v is None:
                return None
            s = str(v or "").strip()
            if not s:
                return None
            try:
                return int(float(s))
            except Exception:
                return None

        wb = load_workbook(str(path), data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return False, "File Excel trống.", []

        headers = [norm_header(h) for h in list(header_row or [])]
        header_to_key_lower = {
            str(k).strip().lower(): v for k, v in header_to_key.items()
        }
        header_to_key_norm = {norm_header_key(k): v for k, v in header_to_key.items()}

        col_keys: list[str | None] = []
        for h in headers:
            key = header_to_key.get(h)
            if key is None:
                key = header_to_key_lower.get(str(h or "").strip().lower())
            if key is None:
                key = header_to_key_norm.get(norm_header_key(h))
            col_keys.append(key)

        unknown_headers = [
            headers[i]
            for i, k in enumerate(col_keys)
            if k is None and str(headers[i] or "").strip()
        ]

        out: list[dict[str, Any]] = []
        for r in rows_iter:
            if r is None:
                continue
            item: dict[str, Any] = {}
            empty = True
            for idx, raw in enumerate(list(r)):
                key = col_keys[idx] if idx < len(col_keys) else None
                if not key or key in {"id"}:
                    continue

                # Preserve STT from Excel when provided, but do not treat it as
                # a data-bearing field for the "empty row" check.
                if key == "stt":
                    item[key] = parse_int(raw)
                    continue
                if raw is not None and str(raw).strip() != "":
                    empty = False

                if key in {
                    "start_date",
                    "date_of_birth",
                    "id_issue_date",
                    "contract1_sign_date",
                    "contract1_expire_date",
                    "contract2_sign_date",
                    "child_dob_1",
                    "child_dob_2",
                    "child_dob_3",
                    "child_dob_4",
                }:
                    # Keep preview display unchanged (e.g. '07/2013' stays '07/2013')
                    item[key] = self._date_value_for_preview(raw)
                elif key == "contract1_signed":
                    # Keep raw text (e.g. "01 năm", "02 năm") for correct display in preview.
                    # Import step will interpret this value using parse_bool/to_bool.
                    if raw is None:
                        item[key] = None
                    else:
                        s_raw = str(raw).strip()
                        item[key] = s_raw if s_raw != "" else None
                elif key == "contract2_indefinite":
                    # Keep raw text for correct preview display (do NOT coerce to bool here).
                    # Import step will interpret this value using parse_bool/to_bool.
                    if raw is None:
                        item[key] = None
                    else:
                        s_raw = str(raw).strip()
                        item[key] = s_raw if s_raw != "" else None
                elif key in {"children_count"}:
                    item[key] = parse_int(raw)
                elif key == "tax_code":
                    if raw is None:
                        item[key] = None
                    elif isinstance(raw, (int, float)):
                        # Excel can treat MST as a number; keep it as text to avoid scientific notation.
                        try:
                            f = float(raw)
                            if f.is_integer():
                                item[key] = str(int(f))
                            else:
                                item[key] = str(raw).strip()
                        except Exception:
                            item[key] = str(raw).strip() or None
                    else:
                        s = str(raw).strip()
                        if not s:
                            item[key] = None
                        elif "e+" in s.lower() or "e-" in s.lower():
                            try:
                                item[key] = str(int(float(s)))
                            except Exception:
                                item[key] = s
                        else:
                            item[key] = s
                else:
                    s = str(raw or "").strip()
                    item[key] = s if s != "" else None

            if empty:
                continue

            # Keep preview values as user entered (only trim outer spaces)
            if item.get("employee_code") is not None:
                item["employee_code"] = (
                    str(item.get("employee_code") or "").strip() or None
                )
            if item.get("full_name") is not None:
                item["full_name"] = str(item.get("full_name") or "").strip() or None

            out.append(item)

        # Add STT for preview
        preview_rows: list[dict[str, Any]] = []
        for idx, it in enumerate(out, start=1):
            row = {"id": None, "stt": idx}
            row.update(it)
            preview_rows.append(row)

        msg = f"Đã đọc {len(preview_rows)} dòng từ Excel."
        if unknown_headers:
            # Keep message short; this helps users fix header typos instead of silently dropping columns.
            sample = ", ".join([str(h) for h in unknown_headers[:6]])
            more = "..." if len(unknown_headers) > 6 else ""
            msg += f" (Không nhận diện {len(unknown_headers)} cột: {sample}{more})"

        return True, msg, preview_rows

    def import_employees_rows(
        self,
        rows: list[dict[str, Any]],
        only_new: bool,
        progress_cb: Callable[[int, bool, str, str], None] | None = None,
        report: list[dict[str, Any]] | None = None,
    ) -> tuple[bool, str]:
        if not rows:
            return False, "Không có dữ liệu để cập nhật."

        # Map title/department name to IDs (best-effort)
        dept_map: dict[str, int] = {}
        for did, dname in self.list_departments_dropdown():
            s = str(dname or "").strip().lower()
            if s:
                dept_map[s] = int(did)

        title_map: dict[str, int] = {}
        for tid, tname in self.list_titles_dropdown():
            s = str(tname or "").strip().lower()
            if s:
                title_map[s] = int(tid)

        def to_bool(v: Any) -> bool:
            parsed = self._parse_bool(v)
            return bool(parsed) if parsed is not None else False

        def norm_tax_code(v: Any) -> str | None:
            if v is None:
                return None
            s = str(v or "").strip()
            if not s:
                return None
            if "e+" in s.lower() or "e-" in s.lower():
                try:
                    return str(int(float(s)))
                except Exception:
                    return s
            return s

        def norm_str(v: Any) -> str | None:
            s = str(v or "").strip()
            return s if s else None

        def norm_status(v: Any) -> str | None:
            s = str(v or "").strip()
            if not s:
                return None
            s_low = s.lower()
            if s_low in {"đi làm", "di lam", "đang làm", "dang lam", "working"}:
                return "Đi làm"
            if s_low in {
                "nghỉ thai sản",
                "nghi thai san",
                "thai sản",
                "thai san",
                "maternity",
                "maternity leave",
                # Common typo seen in user input
                "nghỉ thai sải",
                "nghi thai sai",
            }:
                return "Nghỉ thai sản"
            if s_low in {
                "đã nghỉ việc",
                "da nghi viec",
                "nghỉ việc",
                "nghi viec",
                "resigned",
                "quit",
            }:
                return "Đã nghỉ việc"
            return s

        def norm_payload(it: dict[str, Any]) -> dict[str, Any] | None:
            code = str(it.get("employee_code") or "").strip()
            name = str(it.get("full_name") or "").strip()
            if not code or not name:
                return None
            if not code.isdigit():
                return None
            if len(code) > 5:
                return None

            code = code.zfill(5)

            # Persist Excel order to DB so list can show STT 1..N (top-to-bottom).
            sort_order_val = it.get("stt")
            try:
                sort_order_int = (
                    int(sort_order_val) if sort_order_val is not None else None
                )
            except Exception:
                sort_order_int = None

            title_name = norm_str(it.get("title_name"))
            dept_name = norm_str(it.get("department_name"))

            title_key = str(title_name or "").strip().lower() if title_name else ""
            dept_key = str(dept_name or "").strip().lower() if dept_name else ""

            title_id = title_map.get(title_key) if title_key else None
            dept_id = dept_map.get(dept_key) if dept_key else None

            # Auto-create missing department/title so import does not silently drop values.
            if title_name and not title_id:
                ok, _msg, new_id = self._title_service.create_title(
                    str(title_name),
                    department_id=None,
                )
                if ok and new_id:
                    title_id = int(new_id)
                    title_map[title_key] = int(new_id)

            if dept_name and not dept_id:
                ok, _msg, new_id = self._department_service.create_department(
                    department_name=str(dept_name),
                    parent_id=None,
                    department_note="",
                )
                if ok and new_id:
                    dept_id = int(new_id)
                    dept_map[dept_key] = int(new_id)

            # Keep contract term text (e.g. '01 năm', '02 năm') for display.
            # Source can be either `contract1_term` (future-proof) or the legacy
            # Excel column mapping stored in `contract1_signed`.
            contract1_term = norm_str(it.get("contract1_term")) or norm_str(
                it.get("contract1_signed")
            )

            return {
                "sort_order": sort_order_int,
                "employee_code": code,
                "mcc_code": norm_str(it.get("mcc_code")),
                "full_name": name,
                "name_on_mcc": norm_str(it.get("name_on_mcc")),
                "start_date": self._parse_date_for_db(it.get("start_date")),
                "title_id": title_id,
                "department_id": dept_id,
                "date_of_birth": self._parse_date_for_db(it.get("date_of_birth")),
                "gender": norm_str(it.get("gender")),
                "national_id": norm_str(it.get("national_id")),
                "id_issue_date": self._parse_date_for_db(it.get("id_issue_date")),
                "id_issue_place": norm_str(it.get("id_issue_place")),
                "address": norm_str(it.get("address")),
                "phone": norm_str(it.get("phone")),
                "insurance_no": norm_str(it.get("insurance_no")),
                "tax_code": norm_tax_code(it.get("tax_code")),
                "degree": norm_str(it.get("degree")),
                "major": norm_str(it.get("major")),
                "contract1_signed": bool(
                    to_bool(it.get("contract1_signed")) or (contract1_term is not None)
                ),
                "contract1_term": contract1_term,
                "contract1_no": norm_str(it.get("contract1_no")),
                "contract1_sign_date": self._parse_date_for_db(
                    it.get("contract1_sign_date")
                ),
                "contract1_expire_date": self._parse_date_for_db(
                    it.get("contract1_expire_date")
                ),
                "contract2_indefinite": to_bool(it.get("contract2_indefinite")),
                "contract2_no": norm_str(it.get("contract2_no")),
                "contract2_sign_date": self._parse_date_for_db(
                    it.get("contract2_sign_date")
                ),
                "children_count": it.get("children_count"),
                "child_dob_1": self._parse_date_for_db(it.get("child_dob_1")),
                "child_dob_2": self._parse_date_for_db(it.get("child_dob_2")),
                "child_dob_3": self._parse_date_for_db(it.get("child_dob_3")),
                "child_dob_4": self._parse_date_for_db(it.get("child_dob_4")),
                "employment_status": norm_status(it.get("employment_status")),
                "note": norm_str(it.get("note")),
            }

        def normalize_db_row(db: dict[str, Any]) -> dict[str, Any]:
            def to_iso(v: Any) -> Any:
                if v is None:
                    return None
                try:
                    return v.isoformat()
                except Exception:
                    return v

            return {
                "sort_order": (
                    int(db.get("sort_order"))
                    if db.get("sort_order") is not None
                    and str(db.get("sort_order")) != ""
                    else None
                ),
                "employee_code": str(db.get("employee_code") or "").strip(),
                "mcc_code": (str(db.get("mcc_code") or "").strip() or None),
                "full_name": str(db.get("full_name") or "").strip(),
                "name_on_mcc": (str(db.get("name_on_mcc") or "").strip() or None),
                "start_date": to_iso(db.get("start_date")),
                "title_id": db.get("title_id"),
                "department_id": db.get("department_id"),
                "date_of_birth": to_iso(db.get("date_of_birth")),
                "gender": (str(db.get("gender") or "").strip() or None),
                "national_id": (str(db.get("national_id") or "").strip() or None),
                "id_issue_date": to_iso(db.get("id_issue_date")),
                "id_issue_place": (str(db.get("id_issue_place") or "").strip() or None),
                "address": (str(db.get("address") or "").strip() or None),
                "phone": (str(db.get("phone") or "").strip() or None),
                "insurance_no": (str(db.get("insurance_no") or "").strip() or None),
                "tax_code": (str(db.get("tax_code") or "").strip() or None),
                "degree": (str(db.get("degree") or "").strip() or None),
                "major": (str(db.get("major") or "").strip() or None),
                "contract1_signed": bool(int(db.get("contract1_signed") or 0)),
                "contract1_term": (str(db.get("contract1_term") or "").strip() or None),
                "contract1_no": (str(db.get("contract1_no") or "").strip() or None),
                "contract1_sign_date": to_iso(db.get("contract1_sign_date")),
                "contract1_expire_date": to_iso(db.get("contract1_expire_date")),
                "contract2_indefinite": bool(int(db.get("contract2_indefinite") or 0)),
                "contract2_no": (str(db.get("contract2_no") or "").strip() or None),
                "contract2_sign_date": to_iso(db.get("contract2_sign_date")),
                "children_count": db.get("children_count"),
                "child_dob_1": to_iso(db.get("child_dob_1")),
                "child_dob_2": to_iso(db.get("child_dob_2")),
                "child_dob_3": to_iso(db.get("child_dob_3")),
                "child_dob_4": to_iso(db.get("child_dob_4")),
                "employment_status": (
                    str(db.get("employment_status") or "").strip() or None
                ),
                "note": (str(db.get("note") or "").strip() or None),
            }

        def normalize_payload_for_compare(p: dict[str, Any]) -> dict[str, Any]:
            # Keep only fields that represent employee state.
            outp = dict(p)
            # Normalize ints/bools/strings similarly to normalize_db_row
            try:
                so = outp.get("sort_order")
                outp["sort_order"] = (
                    int(so) if so is not None and str(so) != "" else None
                )
            except Exception:
                outp["sort_order"] = None
            outp["employee_code"] = str(outp.get("employee_code") or "").strip()
            outp["mcc_code"] = str(outp.get("mcc_code") or "").strip() or None
            outp["full_name"] = str(outp.get("full_name") or "").strip()
            outp["name_on_mcc"] = str(outp.get("name_on_mcc") or "").strip() or None
            outp["gender"] = str(outp.get("gender") or "").strip() or None
            outp["national_id"] = str(outp.get("national_id") or "").strip() or None
            outp["id_issue_place"] = (
                str(outp.get("id_issue_place") or "").strip() or None
            )
            outp["address"] = str(outp.get("address") or "").strip() or None
            outp["phone"] = str(outp.get("phone") or "").strip() or None
            outp["insurance_no"] = str(outp.get("insurance_no") or "").strip() or None
            outp["tax_code"] = str(outp.get("tax_code") or "").strip() or None
            outp["degree"] = str(outp.get("degree") or "").strip() or None
            outp["major"] = str(outp.get("major") or "").strip() or None
            outp["contract1_signed"] = bool(outp.get("contract1_signed"))
            outp["contract1_term"] = (
                str(outp.get("contract1_term") or "").strip() or None
            )
            outp["contract1_no"] = str(outp.get("contract1_no") or "").strip() or None
            outp["contract2_indefinite"] = bool(outp.get("contract2_indefinite"))
            outp["contract2_no"] = str(outp.get("contract2_no") or "").strip() or None
            outp["employment_status"] = (
                str(outp.get("employment_status") or "").strip() or None
            )
            outp["note"] = str(outp.get("note") or "").strip() or None
            return outp

        inserted = 0
        updated = 0
        skipped = 0
        invalid = 0
        failed = 0

        def add_report(
            *,
            idx: int,
            code: str,
            name: str,
            result: str,
            action: str,
            message: str,
        ) -> None:
            if report is None:
                return
            report.append(
                {
                    "index": int(idx),
                    "employee_code": str(code or "").strip(),
                    "full_name": str(name or "").strip(),
                    "result": str(result or ""),
                    "action": str(action or ""),
                    "message": str(message or ""),
                }
            )

        # STT/sort_order rules:
        # - Existing employees: never change sort_order (do not create new STT).
        # - New employees: assign a unique sort_order (no duplicates).
        used_sort_orders: set[int] = set()
        max_sort_order: int = 0

        # Always process strictly by the incoming list order (Excel row order / preview order):
        # row 1 -> row 2 -> row 3 ...
        total = len(rows)
        for idx, it in enumerate(rows, start=1):
            raw_code = str(it.get("employee_code") or "").strip()
            raw_name = str(it.get("full_name") or "").strip()

            invalid_reason = ""
            if not raw_code or not raw_name:
                invalid_reason = "Thiếu Mã NV hoặc Họ và tên"
            elif not raw_code.isdigit():
                invalid_reason = "Mã NV phải là số"
            elif len(raw_code) > 5:
                invalid_reason = "Mã NV tối đa 5 ký tự"

            payload = norm_payload(it) if not invalid_reason else None
            code = str((payload or {}).get("employee_code") or raw_code or "").strip()

            if not payload:
                invalid += 1
                reason = invalid_reason or "Dòng không hợp lệ"
                add_report(
                    idx=idx,
                    code=code,
                    name=raw_name,
                    result="INVALID",
                    action="INVALID",
                    message=reason,
                )
                if progress_cb:
                    progress_cb(idx, False, code, reason)
                continue

            try:
                existing = self._repo.get_employee_by_code(code)
                if not existing:
                    # Checkbox meaning: only add new employees when checked.
                    if only_new:
                        # Assign unique STT for new employees. Prefer Excel STT when possible.
                        desired = payload.get("sort_order")
                        try:
                            desired_i = int(desired) if desired is not None else None
                        except Exception:
                            desired_i = None

                        if (
                            desired_i is None
                            or desired_i <= 0
                            or desired_i in used_sort_orders
                        ):
                            max_sort_order = max(
                                max_sort_order, *(used_sort_orders or {0})
                            )
                            desired_i = max_sort_order + 1

                        payload["sort_order"] = desired_i
                        used_sort_orders.add(int(desired_i))
                        max_sort_order = max(max_sort_order, int(desired_i))

                        self._repo.create_employee(payload)
                        inserted += 1
                        add_report(
                            idx=idx,
                            code=code,
                            name=str(payload.get("full_name") or raw_name),
                            result="SUCCESS",
                            action="INSERT",
                            message="Đã thêm",
                        )
                        if progress_cb:
                            progress_cb(idx, True, code, "Đã thêm")
                    else:
                        skipped += 1
                        add_report(
                            idx=idx,
                            code=code,
                            name=raw_name,
                            result="SKIPPED",
                            action="SKIP_NOT_FOUND",
                            message="Bỏ qua (chưa có dữ liệu)",
                        )
                        if progress_cb:
                            progress_cb(idx, True, code, "Bỏ qua (chưa có dữ liệu)")
                    continue

                # Compare all fields: if identical -> skip; if changed -> overwrite.
                db_norm = normalize_db_row(existing)

                # Existing employees: keep current sort_order to avoid creating/overwriting STT.
                payload["sort_order"] = db_norm.get("sort_order")
                if db_norm.get("sort_order") is not None:
                    try:
                        so_i = int(db_norm.get("sort_order") or 0)
                        if so_i > 0:
                            used_sort_orders.add(so_i)
                            max_sort_order = max(max_sort_order, so_i)
                    except Exception:
                        pass

                payload_norm = normalize_payload_for_compare(payload)

                same = True
                for k, v in payload_norm.items():
                    # Only compare keys we also keep in db_norm.
                    if k not in db_norm:
                        continue
                    if db_norm.get(k) != v:
                        same = False
                        break

                if same:
                    skipped += 1
                    add_report(
                        idx=idx,
                        code=code,
                        name=str(payload.get("full_name") or raw_name),
                        result="SKIPPED",
                        action="SKIP_DUPLICATE",
                        message="Bỏ qua (trùng dữ liệu)",
                    )
                    if progress_cb:
                        progress_cb(idx, True, code, "Bỏ qua (trùng dữ liệu)")
                    continue

                self._repo.update_employee(int(existing.get("id")), payload)
                updated += 1
                add_report(
                    idx=idx,
                    code=code,
                    name=str(payload.get("full_name") or raw_name),
                    result="SUCCESS",
                    action="UPDATE",
                    message="Đã cập nhật (ghi đè)",
                )
                if progress_cb:
                    progress_cb(idx, True, code, "Đã cập nhật (ghi đè)")
            except Exception as exc:
                failed += 1
                add_report(
                    idx=idx,
                    code=code,
                    name=raw_name,
                    result="FAILED",
                    action="FAILED",
                    message=str(exc),
                )
                if progress_cb:
                    progress_cb(idx, False, code, str(exc))
                continue

        ok_all = failed == 0
        success = int(inserted) + int(updated)
        return (
            ok_all,
            " | ".join(
                [
                    f"Tổng: {total}",
                    f"Thành công: {success}",
                    f"Thêm mới: {inserted}",
                    f"Cập nhật: {updated}",
                    f"Bỏ qua: {skipped}",
                    f"Lỗi dữ liệu: {invalid}",
                    f"Thất bại: {failed}",
                ]
            ),
        )

    def import_csv(self, file_path: str) -> tuple[bool, str]:
        path = Path(file_path)
        if not path.exists() or path.suffix.lower() != ".csv":
            return False, "Vui lòng chọn file .csv hợp lệ."

        def parse_bool(v: Any) -> bool:
            s = str(v or "").strip().lower()
            return s in {"1", "true", "yes", "y", "x"}

        def parse_int(v: Any) -> int | None:
            s = str(v or "").strip()
            if not s:
                return None
            try:
                return int(s)
            except Exception:
                return None

        def parse_date(v: Any) -> str | None:
            s = str(v or "").strip()
            if not s:
                return None
            # already ISO
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    dt = datetime.strptime(s, fmt)
                    return dt.date().isoformat()
                except Exception:
                    continue
            return None

        items: list[dict[str, Any]] = []
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for r in reader:
                items.append(
                    {
                        "employee_code": (r.get("employee_code") or "").strip(),
                        "full_name": (r.get("full_name") or "").strip(),
                        "start_date": parse_date(r.get("start_date")),
                        "title_id": parse_int(r.get("title_id")),
                        "department_id": parse_int(r.get("department_id")),
                        "date_of_birth": parse_date(r.get("date_of_birth")),
                        "gender": (r.get("gender") or "").strip() or None,
                        "national_id": (r.get("national_id") or "").strip() or None,
                        "id_issue_date": parse_date(r.get("id_issue_date")),
                        "id_issue_place": (r.get("id_issue_place") or "").strip()
                        or None,
                        "address": (r.get("address") or "").strip() or None,
                        "phone": (r.get("phone") or "").strip() or None,
                        "insurance_no": (r.get("insurance_no") or "").strip() or None,
                        "tax_code": (r.get("tax_code") or "").strip() or None,
                        "degree": (r.get("degree") or "").strip() or None,
                        "major": (r.get("major") or "").strip() or None,
                        "contract1_signed": parse_bool(r.get("contract1_signed")),
                        "contract1_no": (r.get("contract1_no") or "").strip() or None,
                        "contract1_sign_date": parse_date(r.get("contract1_sign_date")),
                        "contract1_expire_date": parse_date(
                            r.get("contract1_expire_date")
                        ),
                        "contract2_indefinite": parse_bool(
                            r.get("contract2_indefinite")
                        ),
                        "contract2_no": (r.get("contract2_no") or "").strip() or None,
                        "contract2_sign_date": parse_date(r.get("contract2_sign_date")),
                        "children_count": parse_int(r.get("children_count")),
                        "child_dob_1": parse_date(r.get("child_dob_1")),
                        "child_dob_2": parse_date(r.get("child_dob_2")),
                        "child_dob_3": parse_date(r.get("child_dob_3")),
                        "child_dob_4": parse_date(r.get("child_dob_4")),
                        "note": (r.get("note") or "").strip() or None,
                    }
                )

        affected, skipped = self._repo.upsert_many(items)
        return (
            True,
            f"Đã nhập: {affected} dòng. Bỏ qua: {skipped} dòng (thiếu Mã NV/Họ tên).",
        )

    def create_employee(self, data: dict[str, Any]) -> tuple[bool, str, int | None]:
        code = str(data.get("employee_code") or "").strip()
        name = str(data.get("full_name") or "").strip()

        if not code:
            return False, "Vui lòng nhập Mã NV.", None
        if not code.isdigit():
            return False, "Mã NV chỉ gồm số.", None
        if len(code) > 5:
            return False, "Mã NV tối đa 5 chữ số.", None
        code = code.zfill(5)

        if not name:
            return False, "Vui lòng nhập Họ và tên.", None

        payload = dict(data)
        payload["employee_code"] = code
        payload["full_name"] = name

        # Preserve STT semantics: when sort_order exists and user didn't provide one,
        # assign the next available sort_order so the new employee doesn't appear as STT=1.
        try:
            if payload.get("sort_order") is None:
                next_stt = self._repo.get_next_sort_order()
                if next_stt is not None:
                    payload["sort_order"] = int(next_stt)
        except Exception:
            pass

        try:
            new_id = self._repo.create_employee(payload)
            return True, "Đã thêm nhân viên.", new_id
        except Exception as exc:
            if "1062" in str(exc) or "Duplicate" in str(exc):
                return False, "Mã NV đã tồn tại.", None
            raise

    def update_employee(
        self, employee_id: int, data: dict[str, Any]
    ) -> tuple[bool, str]:
        code = str(data.get("employee_code") or "").strip()
        name = str(data.get("full_name") or "").strip()

        if not code:
            return False, "Vui lòng nhập Mã NV."
        if not code.isdigit():
            return False, "Mã NV chỉ gồm số."
        if len(code) > 5:
            return False, "Mã NV tối đa 5 chữ số."
        code = code.zfill(5)

        if not name:
            return False, "Vui lòng nhập Họ và tên."

        payload = dict(data)
        payload["employee_code"] = code
        payload["full_name"] = name

        # Do not reset STT/contract term when editing from UI.
        try:
            existing = self._repo.get_employee(int(employee_id)) or {}
            if payload.get("sort_order") is None:
                payload["sort_order"] = existing.get("sort_order")
            if payload.get("contract1_term") is None:
                payload["contract1_term"] = existing.get("contract1_term")
        except Exception:
            pass

        try:
            affected = self._repo.update_employee(int(employee_id), payload)
            if affected <= 0:
                return False, "Không tìm thấy nhân viên để cập nhật."
            return True, "Đã cập nhật thông tin."
        except Exception as exc:
            if "1062" in str(exc) or "Duplicate" in str(exc):
                return False, "Mã NV đã tồn tại."
            raise

    def delete_employee(self, employee_id: int) -> tuple[bool, str]:
        affected = self._repo.delete_employee(int(employee_id))
        if affected <= 0:
            return False, "Không tìm thấy nhân viên để xóa."
        try:
            self._repo.resequence_sort_order()
        except Exception:
            pass
        return True, "Đã xóa nhân viên."

    def delete_employees_bulk(
        self,
        employee_ids: list[int],
        progress_cb: Callable[[int, int], None] | None = None,
    ) -> tuple[int, int]:
        ids = [int(i) for i in (employee_ids or []) if int(i) > 0]
        if not ids:
            return 0, 0

        # Unique but keep stable order
        seen: set[int] = set()
        uniq: list[int] = []
        for i in ids:
            if i in seen:
                continue
            seen.add(i)
            uniq.append(i)

        total = len(uniq)
        deleted = 0
        processed = 0

        chunk_size = 200
        for start in range(0, total, chunk_size):
            chunk = uniq[start : start + chunk_size]
            deleted += int(self._repo.delete_employees_bulk(chunk))
            processed = min(total, start + len(chunk))
            if progress_cb:
                progress_cb(processed, total)

        try:
            self._repo.resequence_sort_order()
        except Exception:
            pass

        return deleted, total
